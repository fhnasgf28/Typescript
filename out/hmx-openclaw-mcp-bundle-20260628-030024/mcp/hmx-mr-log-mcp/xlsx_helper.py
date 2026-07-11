#!/usr/bin/env python3
import copy
import json
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

DATE_HINTS = ("date", "tanggal", "created")
HEADER_HINTS = ("date", "tanggal", "mr", "merge", "link", "url", "project", "repo", "title", "judul", "status", "author", "pic", "branch")

def cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)

def normalize(value):
    return " ".join(str(value or "").strip().lower().split())

def normalize_header(value):
    return "".join(ch if ch.isalnum() else " " for ch in normalize(value)).strip()

def detect_header(rows):
    best = {"rowIndex": 0, "score": -1, "row": rows[0] if rows else []}
    for i, row in enumerate(rows[:10]):
        joined = normalize_header(" ".join(row))
        score = sum(1 for hint in HEADER_HINTS if hint in joined) + (sum(1 for c in row if c) / 100.0)
        if score > best["score"]:
            best = {"rowIndex": i, "score": score, "row": row}
    return best

def load_sheet(workbook_path, sheet_name=""):
    wb = load_workbook(workbook_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    return wb, ws

def extract(workbook_path, sheet_name=""):
    wb, ws = load_sheet(workbook_path, sheet_name)
    rows = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        rows.append([cell_value(cell.value) for cell in row])
    return {"sheetName": ws.title, "values": rows}

def copy_row_style(ws, source_row, target_row, max_col):
    if source_row < 1:
        return
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

def insert_row(workbook_path, output_path, row_index_zero, values, sheet_name=""):
    wb, ws = load_sheet(workbook_path, sheet_name)
    target_row = int(row_index_zero) + 1
    max_col = max(len(values), ws.max_column or 1)
    ws.insert_rows(target_row, 1)
    style_source = target_row - 1 if target_row > 1 else target_row + 1
    copy_row_style(ws, style_source, target_row, max_col)
    for col, value in enumerate(values, 1):
        ws.cell(target_row, col).value = value
    wb.save(output_path)
    return {"sheetName": ws.title, "row": target_row, "columns": len(values)}

def main():
    payload = json.load(sys.stdin)
    op = payload.get("op")
    workbook = payload.get("workbook")
    sheet_name = payload.get("sheetName") or ""
    if not workbook:
        raise SystemExit("workbook is required")
    if op == "extract":
        print(json.dumps(extract(workbook, sheet_name)))
        return
    if op == "insert":
        output = payload.get("output")
        if not output:
            raise SystemExit("output is required")
        print(json.dumps(insert_row(workbook, output, payload.get("rowIndexZero"), payload.get("values") or [], sheet_name)))
        return
    raise SystemExit("unknown op")

if __name__ == "__main__":
    main()

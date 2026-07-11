#!/usr/bin/env python3
import copy
import json
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

DATE_HINTS = ("date", "tanggal", "created")
HEADER_HINTS = ("date", "tanggal", "mr", "merge", "link", "url", "project", "repo", "title", "judul", "status", "author", "pic", "branch")

def cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)

def normalize(value):
    return " ".join(str(value or "").strip().lower().split())

def normalize_header(value):
    return "".join(ch if ch.isalnum() else " " for ch in normalize(value)).strip()

def detect_header(rows):
    best = {"rowIndex": 0, "score": -1, "row": rows[0] if rows else []}
    for i, row in enumerate(rows[:10]):
        joined = normalize_header(" ".join(row))
        score = sum(1 for hint in HEADER_HINTS if hint in joined) + (sum(1 for c in row if c) / 100.0)
        if score > best["score"]:
            best = {"rowIndex": i, "score": score, "row": row}
    return best

def load_sheet(workbook_path, sheet_name=""):
    wb = load_workbook(workbook_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    return wb, ws

def extract(workbook_path, sheet_name=""):
    wb, ws = load_sheet(workbook_path, sheet_name)
    rows = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        rows.append([cell_value(cell.value) for cell in row])
    return {"sheetName": ws.title, "values": rows, "maxRow": max_row, "maxCol": max_col}

def last_non_empty_row(ws, max_row, max_col):
    for row_number in range(max_row, 0, -1):
        for col in range(1, max_col + 1):
            value = ws.cell(row_number, col).value
            if value is not None and str(value).strip() != "":
                return row_number
    return 0

def extract_tail(workbook_path, sheet_name="", max_rows=250, header_scan_rows=10):
    wb, ws = load_sheet(workbook_path, sheet_name)
    formatted_max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    max_row = last_non_empty_row(ws, formatted_max_row, max_col) if formatted_max_row and max_col else 0
    max_rows = max(1, int(max_rows or 250))
    header_scan_rows = max(1, int(header_scan_rows or 10))
    header_end = min(header_scan_rows, max_row)
    tail_start = max(header_end + 1, max_row - max_rows + 1)
    rows = []
    seen = set()
    def append_row(row_number):
        if row_number < 1 or row_number > max_row or row_number in seen:
            return
        seen.add(row_number)
        values = [cell_value(ws.cell(row_number, col).value) for col in range(1, max_col + 1)]
        rows.append({"rowIndexZero": row_number - 1, "values": values})
    for row_number in range(1, header_end + 1):
        append_row(row_number)
    for row_number in range(tail_start, max_row + 1):
        append_row(row_number)
    return {"sheetName": ws.title, "rows": rows, "maxRow": max_row, "maxCol": max_col, "tailStartRow": tail_start if max_row else 0}

def copy_row_style(ws, source_row, target_row, max_col):
    if source_row < 1:
        return
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

def apply_insert(ws, row_index_zero, values):
    target_row = int(row_index_zero) + 1
    max_col = max(len(values), ws.max_column or 1)
    ws.insert_rows(target_row, 1)
    style_source = target_row - 1 if target_row > 1 else target_row + 1
    copy_row_style(ws, style_source, target_row, max_col)
    for col, value in enumerate(values, 1):
        ws.cell(target_row, col).value = value
    return {"row": target_row, "columns": len(values)}

def insert_row(workbook_path, output_path, row_index_zero, values, sheet_name=""):
    wb, ws = load_sheet(workbook_path, sheet_name)
    result = apply_insert(ws, row_index_zero, values)
    wb.save(output_path)
    return {"sheetName": ws.title, **result}

def insert_many(workbook_path, output_path, rows, sheet_name=""):
    wb, ws = load_sheet(workbook_path, sheet_name)
    inserted = []
    for item in rows or []:
        result = apply_insert(ws, item.get("rowIndexZero"), item.get("values") or [])
        inserted.append(result)
    wb.save(output_path)
    return {"sheetName": ws.title, "inserted": inserted, "count": len(inserted)}

def update_cells(workbook_path, output_path, cells, sheet_name=""):
    wb, ws = load_sheet(workbook_path, sheet_name)
    updated = []
    for cell in cells or []:
        row = int(cell.get("rowIndexZero")) + 1
        col = int(cell.get("colIndexZero")) + 1
        value = cell.get("value", "")
        ws.cell(row, col).value = value
        updated.append({"row": row, "column": col})
    wb.save(output_path)
    return {"sheetName": ws.title, "updated": updated, "count": len(updated)}

def main():
    payload = json.load(sys.stdin)
    op = payload.get("op")
    workbook = payload.get("workbook")
    sheet_name = payload.get("sheetName") or ""
    if not workbook:
        raise SystemExit("workbook is required")
    if op == "extract":
        print(json.dumps(extract(workbook, sheet_name)))
        return
    if op == "extract_tail":
        print(json.dumps(extract_tail(workbook, sheet_name, payload.get("maxRows") or 250, payload.get("headerScanRows") or 10)))
        return
    if op == "insert":
        output = payload.get("output")
        if not output:
            raise SystemExit("output is required")
        print(json.dumps(insert_row(workbook, output, payload.get("rowIndexZero"), payload.get("values") or [], sheet_name)))
        return
    if op == "insert_many":
        output = payload.get("output")
        if not output:
            raise SystemExit("output is required")
        print(json.dumps(insert_many(workbook, output, payload.get("rows") or [], sheet_name)))
        return
    if op == "update_cells":
        output = payload.get("output")
        if not output:
            raise SystemExit("output is required")
        print(json.dumps(update_cells(workbook, output, payload.get("cells") or [], sheet_name)))
        return
    raise SystemExit("unknown op")

if __name__ == "__main__":
    main()
